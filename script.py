import streamlit as st
import pandas as pd
import os
import re
import urllib.parse
import unicodedata
from io import BytesIO
from openpyxl import load_workbook

# =====================================================================================
#  GARDE-FOU PMU — Validation du naming builder avant upload GCP
#  - Valide la structure + la cohérence (vs onglets de référence Sites / Advertiser / LPS)
#  - Corrige automatiquement les URL (espaces / caractères spéciaux) en préservant les UTM
#  - Affiche un rapport précis (erreurs bloquantes + corrections URL appliquées)
#  - Si aucune erreur bloquante : régénère le MÊME fichier (URLs corrigées) téléchargeable
# =====================================================================================

# ---- Constantes de structure du fichier PMU ----
SHEET_MAIN = "fichier_media"
HEADER_ROW = 12          # ligne des en-têtes de colonnes (0-indexé) -> data à partir de 13
META_COL_LABEL = 0       # colonne A : libellés métadonnées
META_COL_VALUE = 1       # colonne B : valeurs métadonnées
URL_COLUMN = "URL avec les utms"

# ---- Valeurs de référence (doublées en dur comme filet ; les onglets priment si présents) ----
DEFAULT_TRACKING = ["PCC", "RD"]
DEFAULT_FORMAT_TYPE = ["STATIC", "IAB", "VIDEO"]
DEFAULT_PROGRAMMATION = ["Régie", "79"]
DEFAULT_DEVICE = ["MULTIDEVICE", "MOBILE/DESKTOP", "WEB"]
DEFAULT_MODE_ACHAT = ["CPM", "CPM Dyn", "FORFAIT"]

# ---- Métadonnées : ligne attendue (colonne A) -> clé logique ----
META_ROWS = {
    0: "Consultant email",
    7: "Nom de campagne CM",
    11: "Advertiser CM",
}


# =====================================================================================
#  LOGIQUE URL SPÉCIFIQUE PMU
#  Règle : on nettoie UNIQUEMENT les valeurs des paramètres utm_* (avant ET après le #).
#  On ne touche à RIEN d'autre : redirectionUrl (déjà encodé), structure, fragment, etc.
#  Nettoyage des valeurs utm : espaces -> _, accents retirés, caractères spéciaux -> _.
# =====================================================================================
_UTM_KEYS = {"utm_source", "utm_medium", "utm_campaign", "utm_term", "utm_content"}
# Caractères considérés "spéciaux" à remplacer par _ dans une valeur utm
_SPECIALS = set(list('!"#$%&\'()*+,/:;<=?>@[\\]^`{|}~'))


def _strip_accents(s: str) -> str:
    nfkd = unicodedata.normalize("NFKD", s)
    return "".join(ch for ch in nfkd if not unicodedata.combining(ch))


def _sanitize_utm_value(v: str) -> str:
    """Nettoie une valeur utm : accents retirés, espaces et caractères spéciaux -> _."""
    if v is None:
        return v
    s = str(v)
    s = _strip_accents(s)
    s = "".join("_" if (ch in _SPECIALS or ch.isspace()) else ch for ch in s)
    s = re.sub(r"_+", "_", s).strip("_")
    return s


def _clean_utm_in_querystring(qs: str) -> str:
    """
    Prend une chaîne 'a=1&utm_source=x y&b=2', nettoie SEULEMENT les valeurs utm_*,
    laisse tout le reste tel quel (clé et valeur non-utm inchangées).
    Découpe manuelle pour ne rien ré-encoder par accident.
    """
    if not qs:
        return qs
    parts = qs.split("&")
    out = []
    for p in parts:
        if "=" in p:
            k, v = p.split("=", 1)
            if k.lower() in _UTM_KEYS:
                out.append(f"{k}={_sanitize_utm_value(v)}")
            else:
                out.append(p)  # intact
        else:
            out.append(p)  # paramètre sans valeur, intact
    return "&".join(out)


def clean_url(url: str) -> str:
    """
    Nettoyage PMU : on isole la query AVANT le '#' et la query APRÈS le '#',
    et dans chacune on ne nettoie que les valeurs utm_*. Le reste est préservé
    octet pour octet (notamment redirectionUrl=...%3A%2F%2F...).
    """
    if not url or not isinstance(url, str):
        return url

    s = url.strip()

    # 1) Séparer la partie avant '#' et le fragment après '#'
    if "#" in s:
        before_hash, after_hash = s.split("#", 1)
    else:
        before_hash, after_hash = s, None

    # 2) Partie avant '#': structure  base?query
    if "?" in before_hash:
        base, query = before_hash.split("?", 1)
        before_hash = base + "?" + _clean_utm_in_querystring(query)

    # 3) Fragment après '#': structure  /chemin?query  (les utm PMU sont ici)
    if after_hash is not None:
        if "?" in after_hash:
            frag_path, frag_query = after_hash.split("?", 1)
            after_hash = frag_path + "?" + _clean_utm_in_querystring(frag_query)
        s_final = before_hash + "#" + after_hash
    else:
        s_final = before_hash

    return s_final


def describe_url_change(original: str, cleaned: str) -> list:
    """Liste lisible des modifications appliquées à une URL (valeurs utm uniquement)."""
    changes = []
    if not isinstance(original, str) or original == cleaned:
        return changes
    if " " in original:
        changes.append("espace(s) dans une valeur utm remplacé(s) par _")
    # Détection accents dans l'original
    if any(unicodedata.combining(c) for c in unicodedata.normalize("NFKD", original)):
        changes.append("accent(s) retiré(s) dans une valeur utm")
    changes.append("valeur(s) utm nettoyée(s)")
    return list(dict.fromkeys(changes))  # dédoublonne en gardant l'ordre


# =====================================================================================
#  LECTURE DES ONGLETS DE RÉFÉRENCE
# =====================================================================================
def load_reference_values(xls):
    """Charge les valeurs de référence depuis les onglets ; fallback sur les défauts."""
    ref = {
        "tracking": DEFAULT_TRACKING,
        "format_type": DEFAULT_FORMAT_TYPE,
        "programmation": DEFAULT_PROGRAMMATION,
        "device": DEFAULT_DEVICE,
        "mode_achat": DEFAULT_MODE_ACHAT,
        "sites": set(),
        "advertisers": set(),
        "lps_urls": set(),
    }

    def col0_values(sheet):
        try:
            d = pd.read_excel(xls, sheet_name=sheet, header=None)
            return [str(x).strip() for x in d.iloc[:, 0].dropna().tolist()]
        except Exception:
            return []

    if "Type de Tracking" in xls.sheet_names:
        v = col0_values("Type de Tracking")
        if v:
            ref["tracking"] = v
    if "Type Format" in xls.sheet_names:
        v = col0_values("Type Format")
        if v:
            ref["format_type"] = v
    if "Programmation" in xls.sheet_names:
        v = col0_values("Programmation")
        if v:
            ref["programmation"] = v
    if "Device" in xls.sheet_names:
        v = col0_values("Device")
        if v:
            ref["device"] = v
    if "Mode d'Achat" in xls.sheet_names:
        v = col0_values("Mode d'Achat")
        if v:
            ref["mode_achat"] = v

    # Sites : onglet avec header (Site ID / Site Name) -> on garde les Site Name
    if "Sites" in xls.sheet_names:
        try:
            d = pd.read_excel(xls, sheet_name="Sites", header=0)
            name_col = [c for c in d.columns if "name" in str(c).lower()]
            if name_col:
                ref["sites"] = set(str(x).strip() for x in d[name_col[0]].dropna())
        except Exception:
            pass

    # Advertisers : on garde les Advertiser Name
    if "Advertiser" in xls.sheet_names:
        try:
            d = pd.read_excel(xls, sheet_name="Advertiser", header=0)
            name_col = [c for c in d.columns if "name" in str(c).lower()]
            if name_col:
                ref["advertisers"] = set(str(x).strip() for x in d[name_col[0]].dropna())
        except Exception:
            pass

    # LPS : on garde les URL
    if "LPS" in xls.sheet_names:
        try:
            d = pd.read_excel(xls, sheet_name="LPS", header=0)
            url_col = [c for c in d.columns if "url" in str(c).lower()]
            if url_col:
                ref["lps_urls"] = set(str(x).strip() for x in d[url_col[0]].dropna())
        except Exception:
            pass

    return ref


# =====================================================================================
#  VALIDATION
# =====================================================================================
EMAIL_RE = re.compile(r"^[^@\s]+@[^@\s]+\.[^@\s]+$")
DIM_RE = re.compile(r"^\d{1,4}\s*[xX]\s*\d{1,4}$")


def validate_metadata(raw_df, ref):
    """Valide les métadonnées (colonne B). Retourne (liste d'erreurs, email)."""
    errors = []

    # 1) Consultant email : rempli + format valide
    email = raw_df.iloc[0, META_COL_VALUE] if raw_df.shape[0] > 0 else None
    if pd.isna(email) or str(email).strip() == "":
        errors.append("Métadonnée manquante : 'Consultant email' (cellule B1).")
        email = None
    elif not EMAIL_RE.match(str(email).strip()):
        errors.append(f"'Consultant email' invalide : {repr(str(email))} (cellule B1).")
        email = str(email).strip()
    else:
        email = str(email).strip()

    # 2) Nom de campagne CM : rempli
    camp = raw_df.iloc[7, META_COL_VALUE] if raw_df.shape[0] > 7 else None
    if pd.isna(camp) or str(camp).strip() == "":
        errors.append("Métadonnée manquante : 'Nom de campagne CM' (cellule B8).")

    # 3) Advertiser CM : rempli + existe dans l'onglet Advertiser
    adv = raw_df.iloc[11, META_COL_VALUE] if raw_df.shape[0] > 11 else None
    if pd.isna(adv) or str(adv).strip() == "":
        errors.append("Métadonnée manquante : 'Advertiser CM' (cellule B12).")
    elif ref["advertisers"] and str(adv).strip() not in ref["advertisers"]:
        errors.append(f"'Advertiser CM' = {repr(str(adv).strip())} introuvable dans l'onglet Advertiser (cellule B12).")

    return errors, email


EXPECTED_COLUMNS = [
    "Régie", "Programmation", "Support", "Emplacement / Ciblage", "Device",
    "Mode d'Achat", "Type de Tracking", "Type de Format", "Format",
    "Piste Créative", "PLACEMENT NAME DCM", "Créative DCM", "Landing Page",
    "Date de début", "Date de fin", "URL avec les utms",
]

# Colonnes strictement nécessaires aux 7 règles de validation
REQUIRED_COLUMNS = [
    "Support", "Type de Tracking", "Format", "Landing Page", "URL avec les utms",
]


def validate_headers(df):
    """
    Règle 0 : les noms de colonnes doivent être EXACTS (pas d'espace en trop, etc.).
    On vérifie que les colonnes requises existent telles quelles, et on repère
    les colonnes qui ne matchent qu'après strip (= espace parasite).
    """
    errors = []
    actual = list(df.columns)
    actual_stripped = {str(c).strip(): str(c) for c in actual}

    for col in REQUIRED_COLUMNS:
        if col in actual:
            continue
        # la colonne existe mais avec des espaces parasites ?
        if col in actual_stripped:
            raw = actual_stripped[col]
            errors.append(
                f"En-tête incorrect : la colonne {repr(raw)} contient des espaces parasites "
                f"(attendu exactement {repr(col)})."
            )
        else:
            errors.append(f"Colonne requise manquante : {repr(col)}.")
    return errors


def validate_rows(df, ref):
    """Valide chaque ligne sur les règles essentielles. Retourne liste d'erreurs."""
    errors = []

    for idx, row in df.iterrows():
        excel_row = HEADER_ROW + 2 + idx
        prefix = f"Ligne {excel_row}"

        if row.isna().all():
            continue

        # 4) Support : existe dans l'onglet Sites
        if "Support" in df.columns:
            sup = row.get("Support")
            if pd.isna(sup) or str(sup).strip() == "":
                errors.append(f"{prefix} : 'Support' est vide.")
            elif ref["sites"] and str(sup).strip() not in ref["sites"]:
                errors.append(f"{prefix} : 'Support' = {repr(str(sup).strip())} introuvable dans l'onglet Sites.")

        # 5) Landing Page : existe dans l'onglet LPS
        if "Landing Page" in df.columns:
            lp = row.get("Landing Page")
            if pd.isna(lp) or str(lp).strip() == "":
                errors.append(f"{prefix} : 'Landing Page' est vide.")
            elif ref["lps_urls"] and str(lp).strip() not in ref["lps_urls"]:
                errors.append(f"{prefix} : 'Landing Page' = {repr(str(lp).strip())} introuvable dans l'onglet LPS.")

        # 6) Type de Tracking : dans [PCC, RD]
        tracking = str(row.get("Type de Tracking", "")).strip().upper()
        if "Type de Tracking" in df.columns:
            tval = row.get("Type de Tracking")
            if pd.isna(tval) or str(tval).strip() == "":
                errors.append(f"{prefix} : 'Type de Tracking' est vide (attendu : PCC ou RD).")
            elif tracking not in [x.upper() for x in ref["tracking"]]:
                errors.append(f"{prefix} : 'Type de Tracking' = {repr(str(tval).strip())} non autorisé (attendu : {', '.join(ref['tracking'])}).")

        # 7) Si RD -> Format doit être une dimension LxH (PCC : non vérifié, backend force 1x1)
        if "Format" in df.columns and tracking == "RD":
            fmt = row.get("Format")
            if pd.isna(fmt) or str(fmt).strip() == "":
                errors.append(f"{prefix} : 'Format' est vide alors que Type de Tracking = RD (dimension LxH attendue, ex: 300x250).")
            elif not DIM_RE.match(str(fmt).strip()):
                errors.append(f"{prefix} : 'Format' = {repr(str(fmt).strip())} n'est pas une dimension valide pour un RD (ex: 300x250).")

    return errors


def process_urls(df):
    """Corrige les URL de la colonne cible et retourne (df_corrigé, rapport_modifs)."""
    report = []
    if URL_COLUMN not in df.columns:
        return df, report

    df = df.copy()
    for idx, row in df.iterrows():
        excel_row = HEADER_ROW + 2 + idx
        original = row.get(URL_COLUMN)
        if pd.isna(original) or str(original).strip() == "":
            continue
        cleaned = clean_url(str(original))
        if cleaned != str(original):
            changes = describe_url_change(str(original), cleaned)
            report.append({
                "ligne": excel_row,
                "avant": str(original),
                "apres": cleaned,
                "modifs": changes,
            })
            df.at[idx, URL_COLUMN] = cleaned
    return df, report


# =====================================================================================
#  GÉNÉRATION DU FICHIER PROPRE (même structure, URLs corrigées)
# =====================================================================================
def regenerate_clean_file(uploaded_bytes, df_corrected):
    """
    Réécrit le fichier d'origine en ne modifiant QUE la colonne URL de l'onglet
    fichier_media, et en figeant les VALEURS des formules.

    IMPORTANT : les colonnes 'PLACEMENT NAME DCM', 'Créative DCM', 'URL avec les utms'
    sont des formules Excel. openpyxl ne recalcule pas les formules : si on réécrit
    le fichier tel quel, ces cellules ressortent vides côté backend (cm-file-gen),
    ce qui déclenche des erreurs. On ouvre donc en data_only=True pour récupérer les
    dernières valeurs calculées par Excel et les figer en valeurs statiques.
    """
    # data_only=True -> on lit les valeurs calculées (cache Excel), pas les formules
    wb = load_workbook(BytesIO(uploaded_bytes), data_only=True)
    ws = wb[SHEET_MAIN]

    header_row_xl = HEADER_ROW + 1  # openpyxl est 1-indexé
    url_col_idx = None
    for cell in ws[header_row_xl]:
        if cell.value is not None and str(cell.value).strip() == URL_COLUMN:
            url_col_idx = cell.column
            break

    # Mettre à jour la colonne URL avec les valeurs corrigées
    if url_col_idx is not None:
        for i, (_, row) in enumerate(df_corrected.iterrows()):
            xl_row = header_row_xl + 1 + i
            val = row.get(URL_COLUMN)
            if not pd.isna(val):
                ws.cell(row=xl_row, column=url_col_idx, value=str(val))

    out = BytesIO()
    wb.save(out)
    out.seek(0)
    return out


# =====================================================================================
#  RENOMMAGE DES CRÉAS PAR DIMENSION (matching sur Format des lignes RD)
# =====================================================================================
import zipfile
import struct

CREATIVE_NAME_COL = "Créative DCM"
FORMAT_COL = "Format"
TRACKING_COL = "Type de Tracking"
IMG_EXTS = {".jpg", ".jpeg", ".png", ".gif"}
VIDEO_EXTS = {".mp4", ".mov", ".webm", ".avi"}


def get_image_dimensions(file_bytes):
    """Retourne (width, height) d'une image via Pillow, ou None si échec."""
    try:
        from PIL import Image
        img = Image.open(BytesIO(file_bytes))
        return img.size  # (width, height)
    except Exception:
        return None


def get_mp4_dimensions(file_bytes):
    """
    Détection légère des dimensions d'un mp4 en parsant l'atome 'tkhd'.
    Best-effort : renvoie (w, h) ou None (alors renommage manuel).
    """
    try:
        data = file_bytes
        idx = data.find(b'tkhd')
        if idx == -1:
            return None
        # L'atome tkhd : après le nom, version(1)+flags(3), puis selon version
        start = idx + 4
        version = data[start]
        # offset jusqu'à la matrice + width/height (width/height sont les 8 derniers octets de tkhd, en 16.16 fixed)
        # tkhd v0 : taille fixe ; width/height sont les 2 derniers uint32 (fixed point 16.16)
        # On lit les 8 derniers octets de l'atome tkhd
        # Trouver la taille de l'atome (4 octets avant 'tkhd')
        size = struct.unpack('>I', data[idx-4:idx])[0]
        atom_start = idx - 4
        atom_end = atom_start + size
        w_fixed = struct.unpack('>I', data[atom_end-8:atom_end-4])[0]
        h_fixed = struct.unpack('>I', data[atom_end-4:atom_end])[0]
        w = w_fixed >> 16
        h = h_fixed >> 16
        if w > 0 and h > 0:
            return (w, h)
        return None
    except Exception:
        return None


def build_format_index(df):
    """
    Construit un index {(w,h) -> [Créative DCM, ...]} à partir des lignes RD du builder.
    Un même format peut avoir PLUSIEURS créas (une par support) : on les garde toutes.
    Ignore les lignes PCC (pas de créa).
    """
    index = {}
    if CREATIVE_NAME_COL not in df.columns or FORMAT_COL not in df.columns:
        return index

    for _, row in df.iterrows():
        tracking = str(row.get(TRACKING_COL, "")).strip().upper()
        if tracking != "RD":
            continue  # on ignore PCC et autres
        fmt = row.get(FORMAT_COL)
        crea = row.get(CREATIVE_NAME_COL)
        if pd.isna(fmt) or pd.isna(crea):
            continue
        m = re.match(r"^\s*(\d{1,4})\s*[xX]\s*(\d{1,4})\s*$", str(fmt))
        if not m:
            continue
        key = (int(m.group(1)), int(m.group(2)))
        crea_name = str(crea).strip()
        index.setdefault(key, [])
        if crea_name not in index[key]:  # évite les doublons exacts de nom
            index[key].append(crea_name)
    return index


def rename_creatives(files, df):
    """
    files : liste de tuples (filename, bytes)
    Une image d'un format donné est DUPLIQUÉE en une créa par ligne RD de ce format
    (un nom par support).
    Retourne : (zip_bytes, matched, not_found, manual)
      - matched : [(ancien_nom, [nouveaux_noms], dimension)]
      - not_found : [(nom, dimension)] dimension absente du builder
      - manual : [(nom, raison)] dimension non détectée
    """
    index = build_format_index(df)
    matched, not_found, manual = [], [], []

    zip_buffer = BytesIO()
    with zipfile.ZipFile(zip_buffer, "w", zipfile.ZIP_DEFLATED) as zf:
        for fname, fbytes in files:
            ext = os.path.splitext(fname)[1].lower()

            if ext in IMG_EXTS:
                dims = get_image_dimensions(fbytes)
            elif ext in VIDEO_EXTS:
                dims = get_mp4_dimensions(fbytes)
            else:
                manual.append((fname, f"extension {ext} non gérée"))
                continue

            if dims is None:
                manual.append((fname, "dimension non détectée (à renommer à la main)"))
                continue

            if dims in index:
                new_names = []
                for crea_base in index[dims]:
                    new_name = f"{crea_base}{ext}"
                    zf.writestr(new_name, fbytes)   # une copie par support
                    new_names.append(new_name)
                matched.append((fname, new_names, f"{dims[0]}x{dims[1]}"))
            else:
                not_found.append((fname, f"{dims[0]}x{dims[1]}"))

    zip_buffer.seek(0)
    return zip_buffer, matched, not_found, manual


# =====================================================================================
#  INTERFACE STREAMLIT
# =====================================================================================
st.set_page_config(page_title="Garde-fou PMU", page_icon="🛡️", layout="wide")
st.title("🛡️ Garde-fou PMU")

tab_valid, tab_rename = st.tabs(["✅ Validation fichier média", "🖼️ Renommage des créas"])

with tab_valid:
    st.caption("Validez le naming builder PMU avant l'upload GCP. Les URL sont corrigées automatiquement, le reste doit être corrigé à la main si erreur.")

    uploaded_file = st.file_uploader("Dépose le fichier média PMU (naming builder)", type=["xlsx"])

    if uploaded_file is not None:
        uploaded_bytes = uploaded_file.getvalue()
        try:
            xls = pd.ExcelFile(BytesIO(uploaded_bytes))

            if SHEET_MAIN not in xls.sheet_names:
                st.error(f"❌ L'onglet '{SHEET_MAIN}' est introuvable dans le fichier.")
                st.stop()

            raw_df = pd.read_excel(xls, sheet_name=SHEET_MAIN, header=None)
            df = pd.read_excel(xls, sheet_name=SHEET_MAIN, skiprows=HEADER_ROW)
            df = df.dropna(how="all").reset_index(drop=True)

            ref = load_reference_values(xls)

            st.success(f"✅ Fichier chargé — {len(df)} ligne(s) de données détectée(s).")

            if st.button("🔍 Valider le fichier"):
                header_errors = validate_headers(df)
                meta_errors, _ = validate_metadata(raw_df, ref)
                row_errors = validate_rows(df, ref)
                df_corrected, url_report = process_urls(df)

                all_errors = header_errors + meta_errors + row_errors

                # ---- Rapport des corrections URL (informatif, non bloquant) ----
                if url_report:
                    st.subheader("🔧 Corrections d'URL appliquées")
                    st.caption("Ces corrections sont automatiques. Vérifie qu'elles te conviennent.")
                    for r in url_report:
                        with st.expander(f"Ligne {r['ligne']} — {', '.join(r['modifs'])}"):
                            st.markdown("**Avant :**")
                            st.code(r["avant"], language="text")
                            st.markdown("**Après :**")
                            st.code(r["apres"], language="text")
                else:
                    st.info("ℹ️ Aucune correction d'URL nécessaire — les URL étaient déjà propres.")

                # ---- Erreurs bloquantes ----
                if all_errors:
                    st.subheader("❌ Erreurs à corriger")
                    st.error(f"{len(all_errors)} erreur(s) bloquante(s). Corrige le fichier source puis recharge-le.")
                    for e in all_errors:
                        st.markdown(f"- {e}")
                    st.warning("🚫 Aucun fichier propre généré tant que les erreurs ne sont pas corrigées.")
                else:
                    st.subheader("✅ Validation réussie")
                    st.success("Aucune erreur bloquante. Le fichier propre est prêt à être uploadé sur GCP.")
                    clean_io = regenerate_clean_file(uploaded_bytes, df_corrected)

                    # Nom du fichier basé sur le Nom de campagne CM (cellule B8)
                    camp_name = raw_df.iloc[7, META_COL_VALUE] if raw_df.shape[0] > 7 else None
                    if pd.isna(camp_name) or str(camp_name).strip() == "":
                        out_name = uploaded_file.name.replace(".xlsx", "_VALIDE.xlsx")
                    else:
                        safe = re.sub(r'[\\/:*?"<>|]+', "_", str(camp_name).strip())
                        safe = re.sub(r"\s+", "_", safe)
                        out_name = f"{safe}.xlsx"

                    st.download_button(
                        label="📥 Télécharger le fichier propre",
                        data=clean_io,
                        file_name=out_name,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )

        except Exception as e:
            st.error(f"❌ Erreur lors du traitement du fichier : {e}")


with tab_rename:
    st.caption("Renomme automatiquement les créas selon la colonne 'Créative DCM', en détectant la dimension de chaque image. Les lignes PCC sont ignorées (pas de créa).")

    col1, col2 = st.columns(2)
    with col1:
        builder_file = st.file_uploader("1) Fichier média (builder validé)", type=["xlsx"], key="rename_builder")
    with col2:
        crea_files = st.file_uploader(
            "2) Créas reçues (jpg, png, gif, mp4)",
            type=["jpg", "jpeg", "png", "gif", "mp4", "mov", "webm", "avi"],
            accept_multiple_files=True,
            key="rename_creas",
        )

    if builder_file is not None and crea_files:
        try:
            xls_b = pd.ExcelFile(BytesIO(builder_file.getvalue()))
            if SHEET_MAIN not in xls_b.sheet_names:
                st.error(f"❌ L'onglet '{SHEET_MAIN}' est introuvable dans le builder.")
                st.stop()
            df_b = pd.read_excel(xls_b, sheet_name=SHEET_MAIN, skiprows=HEADER_ROW)
            df_b = df_b.dropna(how="all").reset_index(drop=True)

            if st.button("🔁 Renommer les créas"):
                files = [(f.name, f.getvalue()) for f in crea_files]
                zip_buffer, matched, not_found, manual = rename_creatives(files, df_b)

                if matched:
                    total_out = sum(len(nn) for _, nn, _ in matched)
                    st.subheader(f"✅ {len(matched)} créa(s) reçue(s) → {total_out} fichier(s) généré(s)")
                    for old, new_names, dim in matched:
                        st.markdown(f"- `{old}` ({dim}) → **{len(new_names)}** copie(s) :")
                        for nn in new_names:
                            st.markdown(f"    - {nn}")

                if not_found:
                    st.subheader("❌ Dimensions absentes du fichier Excel")
                    st.error("Ces créas ont une dimension qui n'existe dans aucune ligne RD du builder (vérifie s'il manque une ligne) :")
                    for name, dim in not_found:
                        st.markdown(f"- `{name}` → dimension **{dim}** introuvable dans l'Excel")

                if manual:
                    st.subheader("✋ À renommer à la main")
                    for name, reason in manual:
                        st.markdown(f"- `{name}` → {reason}")

                if matched:
                    st.download_button(
                        label="📥 Télécharger les créas renommées (ZIP)",
                        data=zip_buffer,
                        file_name="creas_renommees.zip",
                        mime="application/zip",
                    )
                else:
                    st.info("Aucune créa renommée — rien à télécharger.")

        except Exception as e:
            st.error(f"❌ Erreur lors du renommage : {e}")
